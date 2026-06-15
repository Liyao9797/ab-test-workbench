import json
import re
import shutil
from pathlib import Path
from uuid import uuid4

import pandas as pd
from fastapi import UploadFile
from openpyxl import load_workbook

from app.core.paths import STORAGE_ROOT, ensure_storage_dirs


def save_upload(file: UploadFile) -> dict[str, object]:
    ensure_storage_dirs()
    upload_id = f"upl_{uuid4().hex[:12]}"
    upload_dir = STORAGE_ROOT / "uploads" / upload_id
    upload_dir.mkdir(parents=True, exist_ok=False)

    original_path = upload_dir / "original.xlsx"
    with original_path.open("wb") as output:
        shutil.copyfileobj(file.file, output)

    workbook = load_workbook(original_path, read_only=True, data_only=True)
    sheets = workbook.sheetnames
    workbook.close()
    if not sheets:
        raise ValueError("Workbook has no sheets.")

    metadata = {
        "upload_id": upload_id,
        "filename": file.filename,
        "file_size": original_path.stat().st_size,
        "sheets": sheets,
        "selected_sheet": sheets[0],
        "path": str(original_path),
    }
    _write_json(upload_dir / "metadata.json", metadata)
    return metadata


def detect_headers(upload_id: str, sheet_name: str, header_row: int | None, preview_rows: int) -> dict[str, object]:
    upload_dir = STORAGE_ROOT / "uploads" / upload_id
    workbook_path = upload_dir / "original.xlsx"
    if not workbook_path.exists():
        raise FileNotFoundError(f"Upload not found: {upload_id}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    detected_header_row = header_row or _detect_header_row(sheet)
    raw_headers = [cell.value for cell in sheet[detected_header_row]]
    columns = [_normalize_header(value, index) for index, value in enumerate(raw_headers)]

    preview: list[dict[str, object]] = []
    samples_by_column: dict[str, list[object]] = {column: [] for column in columns}
    for row in sheet.iter_rows(min_row=detected_header_row + 1, max_row=detected_header_row + preview_rows, values_only=True):
        record: dict[str, object] = {}
        for index, column in enumerate(columns):
            value = row[index] if index < len(row) else None
            safe_value = _safe_cell_value(value)
            record[column] = safe_value
            if safe_value is not None and len(samples_by_column[column]) < 5:
                samples_by_column[column].append(safe_value)
        if any(value is not None for value in record.values()):
            preview.append(record)

    row_count = max(sheet.max_row - detected_header_row, 0)
    workbook.close()

    response = {
        "upload_id": upload_id,
        "sheet_name": sheet_name,
        "detected_header_row": detected_header_row,
        "columns": [
            {
                "name": column,
                "index": index,
                "dtype": _guess_dtype(samples_by_column[column]),
                "sample_values": samples_by_column[column],
            }
            for index, column in enumerate(columns)
        ],
        "row_count": row_count,
        "preview": preview,
    }
    _write_json(upload_dir / "detected_headers.json", response)
    return response


def create_chi_square_demo_upload() -> dict[str, object]:
    ensure_storage_dirs()
    upload_id = f"upl_{uuid4().hex[:12]}"
    upload_dir = STORAGE_ROOT / "uploads" / upload_id
    upload_dir.mkdir(parents=True, exist_ok=False)
    original_path = upload_dir / "original.xlsx"

    rows = []
    channels = ["organic", "ads_meta", "ads_google", "ads_tiktok"]
    platforms = ["ios", "android"]
    outcome_a = ["low_engagement"] * 96 + ["medium_engagement"] * 144 + ["high_engagement"] * 80
    outcome_b = ["low_engagement"] * 64 + ["medium_engagement"] * 128 + ["high_engagement"] * 128
    reward_a = ["coins"] * 128 + ["energy"] * 112 + ["skin"] * 80
    reward_b = ["coins"] * 96 + ["energy"] * 104 + ["skin"] * 120

    uid_seed = 2800000000
    for group, outcomes, rewards in [("group_a", outcome_a, reward_a), ("group_b", outcome_b, reward_b)]:
        for index, (outcome, reward) in enumerate(zip(outcomes, rewards)):
            rows.append(
                {
                    "uid": str(uid_seed + len(rows)),
                    "group_id": group,
                    "is_new_user": 1 if index % 2 == 0 else 0,
                    "channel": channels[index % len(channels)],
                    "platform": platforms[index % len(platforms)],
                    "outcome_category": outcome,
                    "reward_preference": reward,
                    "session_type": "short" if index % 3 == 0 else ("medium" if index % 3 == 1 else "long"),
                    "feature_clicked": 1 if (group == "group_b" and index % 10 < 6) or (group == "group_a" and index % 10 < 5) else 0,
                    "login_days": 2 + (index % 5) + (1 if group == "group_b" and index % 4 == 0 else 0),
                }
            )

    frame = pd.DataFrame(rows)
    summary = (
        frame.groupby(["group_id", "outcome_category"])
        .size()
        .reset_index(name="user_count")
    )
    with pd.ExcelWriter(original_path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="chi_square_demo", index=False)
        summary.to_excel(writer, sheet_name="category_summary", index=False)

    metadata = {
        "upload_id": upload_id,
        "filename": "chi_square_demo_ab_test.xlsx",
        "file_size": original_path.stat().st_size,
        "sheets": ["chi_square_demo", "category_summary"],
        "selected_sheet": "chi_square_demo",
        "path": str(original_path),
    }
    _write_json(upload_dir / "metadata.json", metadata)
    return metadata


def _detect_header_row(sheet) -> int:
    for row_index in range(1, min(sheet.max_row, 10) + 1):
        values = [cell.value for cell in sheet[row_index]]
        non_empty = [value for value in values if value is not None and str(value).strip()]
        if len(non_empty) >= 2:
            return row_index
    return 1


def _normalize_header(value: object, index: int) -> str:
    fallback = f"column_{index + 1}"
    if value is None:
        return fallback
    text = str(value).strip()
    if not text:
        return fallback
    return re.sub(r"\s+", "_", text)


def _guess_dtype(values: list[object]) -> str:
    if not values:
        return "empty"
    if all(isinstance(value, bool) for value in values):
        return "binary"
    if all(isinstance(value, (int, float)) and value in (0, 1) for value in values):
        return "binary"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
        return "number"
    return "string"


def _safe_cell_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _write_json(path: Path, payload: dict[str, object]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
